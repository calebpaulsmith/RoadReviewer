#!/usr/bin/env python3
"""Generate the Survey123 companion for the Cat C Road/LWC filler.

Emits two files that must stay in sync, which is why one script writes both:

- catc_road_lwc_xlsform.xlsx — an XLSForm to publish with Survey123 Connect
  (or the web designer's XLSForm import). Field crews collect in the
  Survey123 app; every submission lands in the hosted feature layer.
- survey123_map.json — Survey123 field name -> PDF AcroForm field name
  mapping, inlined into the web app by build.py so its "Import Survey123
  CSV" button can populate the form from a Data-tab CSV export.

Survey123 names are snake_case (AGOL rejects spaces/#/% in field names);
the PDF names keep FEMA's original spellings, typos included.

Run: python3 gen_xlsform.py   (writes into this directory)
"""
import json
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Font

HERE = pathlib.Path(__file__).resolve().parent

YNU = "yes_no_unsure"

# ---- the single source of truth ------------------------------------------
# (kind, survey_name, label, pdf_target, extra)
# kind: text | note_text (multiline) | date | decimal | choice | multi
# choice/multi pdf_target: {choice_name: pdf_checkbox}
TEXT = [
    ("applicant", "Applicant", "Applicant3"),
    ("pa_id", "PA ID #", "PA ID3"),
    ("app_rep", "Applicant Representative", "App Rep3"),
    ("app_rep_title", "Representative Title", "App Rep Title3"),
    ("si_name", "Site Inspector Name", "SI Name3"),
    ("work_order", "Work Order #", "Work Order3"),
    ("damage_num", "Damage #", "Damage3"),
    ("year_built", "Year Built", "Year Built3"),
    ("lanes", "Number of Lanes", "# of Lanes3"),
    ("length", "Length", "Length3"),
    ("width", "Width", "Width3"),
    ("xsec_easement", "Cross-section: Easement", "Text11"),
    ("xsec_surface", "Cross-section: Surface", "Text12"),
    ("xsec_base", "Cross-section: Base", "Text13"),
    ("xsec_subbase", "Cross-section: Subbase", "Text14"),
    ("gps_end_lat", "GPS End Latitude", "GPS Latitude End3"),
    ("gps_end_lon", "GPS End Longitude", "GPS Longitude End3"),
    ("mit_other", "Other mitigation (specify)", "Other Mitigation7"),
]
AREAS = [
    ("description", "Facility description / notes", "Notes3"),
    ("photo_notes", "Photo page notes", "Notes4"),
    ("mit_cause", "1. Specific cause of damage", "Mitigation Considerations3"),
    ("mit_q3_comments", "3. Comments", "Comment27"),
    ("mit_q4_comments", "4. Comments", "Comment28"),
    ("mit_q5_comments", "5. Comments", "Comment29"),
    ("mit_notes", "Additional mitigation notes / scope", "Comment30"),
    ("ins_q1_comments", "Insurance comments", "Comment31"),
    ("ehp_q1_comments", "EHP 1 comments", "Comment32"),
    ("ehp_q2_comments", "EHP 2 comments", "Comment33"),
    ("ehp_q3_comments", "EHP 3 comments", "Comment34"),
    ("ehp_q4_comments", "EHP 4 comments", "Comment35"),
    ("ehp_q5_comments", "EHP 5 comments", "Comment36"),
    ("ehp_q6_comments", "EHP 6 comments", "Comment37"),
    ("ehp_q7_comments", "EHP 7 comments", "Comment38"),
    ("ehp_q8_comments", "EHP 8 comments", "Comment39"),
    ("additional_notes", "Additional notes / comments", "Additional Notes/Comments3"),
]
DATES = [
    ("si_date", "Site Inspection Date", "SI Date3"),
    ("date_damaged", "Date Damaged", "Date Damaged3"),
]
GPS = [  # calculated from the map point; still export to CSV
    ("gps_start_lat", "GPS Start Latitude", "GPS Latitude Start3"),
    ("gps_start_lon", "GPS Start Longitude", "GPS Longitude Start3"),
]
CHOICES = [
    ("facility", "Facility", "facility_list",
     {"road": "Check Box147", "lwc": "Check Box148"}),
    ("age", "Age of Facility", "age_list",
     {"exact": "Check Box149", "approximate": "Check Box150"}),
    ("legal_responsibility", "Legal Responsibility", "yes_no",
     {"yes": "Check Box151", "no": "Check Box152"}),
    ("mit_q3", "3. Applicant plans additional protective work?", YNU,
     {"yes": "Check Box177", "no": "Check Box178", "unsure": "Check Box179"}),
    ("mit_q4", "4. Applicant will provide a mitigation proposal?", YNU,
     {"yes": "Check Box180", "no": "Check Box181", "unsure": "Check Box182"}),
    ("mit_q5", "5. Applicant wants FEMA to prepare a proposal?", YNU,
     {"yes": "Check Box183", "no": "Check Box184", "unsure": "Check Box185"}),
    ("ins_q1", "Insurance coverage / insurable risk?", YNU,
     {"yes": "Check Box186", "no": "Check Box187", "unsure": "Check Box188"}),
    ("ehp_q1", "EHP 1. Floodplain / wetland / within 200 ft of waterway?", YNU,
     {"yes": "Check Box189", "no": "Check Box190", "unsure": "Check Box191"}),
    ("ehp_q2", "EHP 2. Coastal Barrier Resource System / protected area?", YNU,
     {"yes": "Check Box192", "no": "Check Box193", "unsure": "Check Box194"}),
    ("ehp_q3", "EHP 3. Repairs change pre-disaster conditions?", YNU,
     {"yes": "Check Box195", "no": "Check Box196", "unsure": "Check Box197"}),
    ("ehp_q4", "EHP 4. Historic register / landmark / older than 45 years?", YNU,
     {"yes": "Check Box198", "no": "Check Box199", "unsure": "Check Box200"}),
    ("ehp_q5", "EHP 5. Large undeveloped areas on or near site?", YNU,
     {"yes": "Check Box201", "no": "Check Box202", "unsure": "Check Box203"}),
    ("ehp_q6", "EHP 6. Hazardous materials at or adjacent?", YNU,
     {"yes": "Check Box204", "no": "Check Box205", "unsure": "Check Box206"}),
    ("ehp_q7", "EHP 7. Other environmental / controversial issues?", YNU,
     {"yes": "Check Box207", "no": "Check Box208", "unsure": "Check Box209"}),
    ("ehp_q8", "EHP 8. Known endangered species in work area?", YNU,
     {"yes": "Check Box210", "no": "Check Box211", "unsure": "Check Box212"}),
]
MULTIS = [
    ("road_type", "Road Type", "road_type_list",
     {"asphalt": "Check Box153", "concrete": "Check Box154",
      "composite": "Check Box155", "chip_seal": "Check Box156",
      "dirt": "Check Box157", "gravel": "Check Box158"}),
    ("mit_options", "2. What can prevent future damage?", "mitigation_list",
     {"m1": "Check Box171", "m2": "Check Box172", "m3": "Check Box173",
      "m4": "Check Box174", "m5": "Check Box175", "m6": "Check Box176"}),
]
# component-damage site blocks; PDF names keep FEMA's typos
SITES = [
    (1, "Site9", "Damage Components9", "Location Address9", "Damage Dimensions9",
     "Method of Repair9", "Cause of Damage9", "Quantity9", "Units9", "%Complete9",
     {"fa": "Check Box159", "ctr": "Check Box160", "both": "Check Box161"}),
    (2, "Site10", "Damage Components10", "Location Address10", "Damage Dimensions10",
     "Method of Repair10", "Cause of Damage10", "Quantity10", "Units10", "% Complete 10",
     {"fa": "Check Box162", "ctr": "Check Box163", "both": "Check Box164"}),
    (3, "Site11", "Damage Components11", "Location Addres11", "Damage Dimensions11",
     "Method of Repair11", "Cause of Damage11", "Quantity 11", "Units11", "% Complete11",
     {"fa": "Check Box165", "ctr": "Check Box166", "both": "Check Box167"}),
    (4, "Site12", "Damage Components12", "Location Address12", "Damage Dimensions12",
     "Method of Repair12", "Cause of Damage12", "Quantity12", "Units12", "% Complete12",
     {"fa": "Check Box168", "ctr": "Check Box169", "both": "Check Box170"}),
]

CHOICE_LISTS = {
    "yes_no_unsure": [("yes", "Yes"), ("no", "No"), ("unsure", "Unsure")],
    "yes_no": [("yes", "Yes"), ("no", "No")],
    "facility_list": [("road", "Road"), ("lwc", "Low Water Crossing")],
    "age_list": [("exact", "Exact"), ("approximate", "Approximate")],
    "road_type_list": [("asphalt", "Asphalt"), ("concrete", "Concrete"),
                       ("composite", "Composite"), ("chip_seal", "Chip & Seal"),
                       ("dirt", "Dirt"), ("gravel", "Gravel")],
    "mitigation_list": [
        ("m1", "Deeper concrete toes along LWC edges"),
        ("m2", "Anchor LWC toes with rebar into bedrock"),
        ("m3", "Reinforced concrete instead of grouted riprap on slopes"),
        ("m4", "Replace LWC with a bridge (needs H&H study)"),
        ("m5", "Increase steel reinforcing in slab and toes"),
        ("m6", "Other mitigation (specify)")],
    "work_by_list": [("fa", "FA (force account)"), ("ctr", "CTR (contract)"),
                     ("both", "Both")],
}


def build_map():
    m = {"text": {}, "date": {}, "choice": {}, "multi": {}}
    for name, _label, pdf in TEXT + AREAS + GPS:
        m["text"][name] = pdf
    for name, _label, pdf in DATES:
        m["date"][name] = pdf
    for name, _label, _list, targets in CHOICES:
        m["choice"][name] = targets
    for name, _label, _list, targets in MULTIS:
        m["multi"][name] = targets
    for (n, site, comp, loc, dim, mor, cause, qty, units, pct, work) in SITES:
        p = f"site{n}_"
        m["text"].update({p + "num": site, p + "component": comp,
                          p + "location": loc, p + "dimensions": dim,
                          p + "repair": mor, p + "cause": cause,
                          p + "qty": qty, p + "units": units, p + "pct": pct})
        m["choice"][p + "work_by"] = work
    return m


def build_xlsform():
    wb = Workbook()
    bold = Font(name="Arial", bold=True)
    plain = Font(name="Arial")

    sv = wb.active
    sv.title = "survey"
    headers = ["type", "name", "label", "hint", "appearance", "calculation", "required"]
    sv.append(headers)

    def row(type_, name="", label="", hint="", appearance="", calculation="", required=""):
        sv.append([type_, name, label, hint, appearance, calculation, required])

    row("begin group", "g_header", "Applicant & inspection")
    for name, label, _ in TEXT[:7]:
        row("text", name, label, required="yes" if name in ("applicant", "damage_num") else "")
    for name, label, _ in DATES[:1]:
        row("date", name, label)
    row("end group")

    row("begin group", "g_location", "Location")
    row("geopoint", "location", "Site location (start point)",
        "Drop the pin at the start of the damaged section")
    for name, label, _ in GPS:
        axis = "y" if name.endswith("lat") else "x"
        row("calculate", name, label,
            calculation=f'pulldata("@geopoint", ${{location}}, "{axis}")')
    row("text", "gps_end_lat", "GPS End Latitude",
        "Needed for facilities longer than 200 linear ft")
    row("text", "gps_end_lon", "GPS End Longitude")
    row("end group")

    row("begin group", "g_facility", "Facility details")
    row("date", "date_damaged", "Date Damaged")
    row("select_one facility_list", "facility", "Facility")
    row("select_one age_list", "age", "Age of Facility")
    row("text", "year_built", "Year Built")
    row("text", "lanes", "Number of Lanes")
    row("select_one yes_no", "legal_responsibility", "Legal Responsibility")
    row("select_multiple road_type_list", "road_type", "Road Type")
    row("text", "length", "Length")
    row("text", "width", "Width")
    row("end group")

    row("begin group", "g_description", "Description & sketch")
    row("text", "description", "Facility description / notes",
        "Pre-disaster design, function, capacity, dimensions, footprint",
        appearance="multiline")
    for name, label, _ in TEXT[11:15]:
        row("text", name, label)
    row("image", "sketch", "Sketch (photo of a hand sketch is fine)")
    row("end group")

    for (n, *_rest) in SITES:
        p = f"site{n}_"
        row("begin group", f"g_site{n}", f"Damage site {n}")
        row("text", p + "num", "Site #")
        row("text", p + "component", "Damage component",
            "Material / model / type / capacity")
        row("text", p + "location", "Location", "Address / GPS / begin-end")
        row("text", p + "dimensions", "Damage dimensions", "L x W x D/L x Dia")
        row("text", p + "repair", "Method of repair",
            "Change in design, materials, size, capacity", appearance="multiline")
        row("text", p + "cause", "Cause of damage", "e.g. 1 - Surface water flooding")
        row("select_one work_by_list", p + "work_by", "Work performed by")
        row("text", p + "qty", "Quantity")
        row("text", p + "units", "Units")
        row("text", p + "pct", "% Complete")
        row("end group")

    row("begin group", "g_photos", "Photos")
    for i in (1, 2, 3, 4):
        row("image", f"photo{i}", f"Site photo {i}")
    row("text", "photo_notes", "Photo notes", appearance="multiline")
    row("end group")

    row("begin group", "g_mitigation", "Mitigation considerations")
    row("text", "mit_cause", "1. Specific cause of damage", appearance="multiline")
    row("select_multiple mitigation_list", "mit_options",
        "2. What can prevent future damage? (check all of interest)")
    row("text", "mit_other", "Other mitigation (specify)")
    for name, label, _list, _t in CHOICES[3:6]:
        row(f"select_one {YNU}", name, label)
        row("text", name + "_comments", "Comments", appearance="multiline")
    row("text", "mit_notes", "Additional mitigation notes / scope",
        appearance="multiline")
    row("end group")

    row("begin group", "g_insurance", "Insurance")
    row(f"select_one {YNU}", "ins_q1", CHOICES[6][1])
    row("text", "ins_q1_comments", "Comments", appearance="multiline")
    row("end group")

    row("begin group", "g_ehp", "Environmental & Historic Preservation")
    for name, label, _list, _t in CHOICES[7:]:
        row(f"select_one {YNU}", name, label)
        row("text", name + "_comments", "Comments", appearance="multiline")
    row("end group")

    row("text", "additional_notes", "Additional notes / comments",
        appearance="multiline")
    row("image", "signature", "Applicant Representative signature",
        appearance="signature")

    ch = wb.create_sheet("choices")
    ch.append(["list_name", "name", "label"])
    for list_name, options in CHOICE_LISTS.items():
        for name, label in options:
            ch.append([list_name, name, label])

    st = wb.create_sheet("settings")
    st.append(["form_title", "form_id", "instance_name"])
    st.append(["Cat C Road-LWC Site Inspection", "catc_road_lwc",
               "concat(${damage_num}, ' - ', ${applicant})"])

    for ws in (sv, ch, st):
        for cell in ws[1]:
            cell.font = bold
        for r in ws.iter_rows(min_row=2):
            for cell in r:
                cell.font = plain
        for col, width in zip("ABCDEFG", (26, 22, 58, 44, 12, 44, 9)):
            ws.column_dimensions[col].width = width

    return wb


if __name__ == "__main__":
    mapping = build_map()
    (HERE / "survey123_map.json").write_text(json.dumps(mapping, indent=1))
    wb = build_xlsform()
    wb.save(HERE / "catc_road_lwc_xlsform.xlsx")
    n = sum(len(v) for v in mapping.values())
    print(f"wrote catc_road_lwc_xlsform.xlsx and survey123_map.json ({n} mapped fields)")
